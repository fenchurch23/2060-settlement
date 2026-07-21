#!/usr/bin/env python3
"""2060 Settlement Atlas — data-driven entry generator + CW-column adapter.
Reads the newest master/c-file in /mnt/project (prefers vNNNc over vNNN, higher NNN wins).
Data fields come from research columns; reader copy comes from CW columns 82-94 (1-idx)/81-93 (0-idx).
The per-region content file supplies only the artisanal layer (hero SVG, marker, hero-adjacent
labels, season labels, cultural_draw/legacy overrides). Content-file prose is a fallback for
regions the copywriter hasn't reached yet.
"""
GENERATOR_VERSION = "1.1 (2026-07-14: energy stat box shows keyword summary; full KEY TRANSITION moved to Key Facts energy row — authorized 2026-07-14)"
import re, sys, glob, os, importlib.util
from openpyxl import load_workbook

def newest_workbook():
    """Read target = the current canon. Highest version number wins; a c-suffix breaks
    ties at the SAME number (vNNNc > vNNN). Works whether or not the canon carries a
    'c' (v130 carries copy without one). NOTE: this assumes the highest-numbered file is
    always the copy canon — true under the merge_cw workflow. If a bare master were ever
    saved at a higher number than the newest copy file, this would read it; the safeguard
    is the naming convention, not the code."""
    cands = glob.glob('/mnt/project/2060_Settlement_Atlas_v*.xlsx')
    if not cands: raise SystemExit('no workbook found in /mnt/project')
    def ver(p):
        m = re.search(r'v(\d+)(c?)\.xlsx$', p)
        return (int(m.group(1)), 1 if m.group(2) == 'c' else 0) if m else (0, 0)
    return max(cands, key=ver)

WB = newest_workbook()

def load_region(region_id):
    ws = load_workbook(WB, read_only=True, data_only=True)['Region Catalog']
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if r[56] == region_id: return r
    raise SystemExit(f'Region ID {region_id} not found in {WB}')

def num(v):
    try: return float(str(v).replace('M','').replace(',','').strip())
    except: return None

def after(label, s, end='|'):
    if not s: return ''
    m = re.search(re.escape(label)+r'\s*([^'+re.escape(end)+r'\n]+)', str(s))
    return m.group(1).strip() if m else ''

def cw_prose(r):
    if r[81] in (None, ''): return {}
    g = lambda c: (str(r[c]).strip() if r[c] not in (None,'') else '')
    split4 = lambda c: [x.strip() for x in g(c).split('|') if x.strip()]
    return {'subhead':g(81), 'lede':[p for p in (g(82),g(83)) if p], 'holding':g(84),
            'living':split4(85), 'table_prose':g(86), 'free_hours':split4(87), 'ledger':g(88),
            'captions':[g(89),g(90)], 'cta_big':g(91), 'cta_small':g(92), 'status':g(93)}

def water(r):  return (after('RATING:',r[8]) or '—').strip(), (after('2060 RISK:',r[9]) or after('RISK:',r[9]) or '—').strip()
ENERGY_KW = ['offshore wind','wind','solar','geothermal','micro-hydro','hydro','biomass','biogas',
             'nuclear','tidal','wave','wood','desal','storage']
def energy(r):
    """rating for the stat-box headline; a keyword summary for the stat-box line
    (v1.1 — long KEY TRANSITION paragraphs overflowed the box); the FULL detail
    for the Key Facts energy row."""
    rating=(after('RATING:',r[49]) or '—').strip()
    full=(after('KEY TRANSITION:',r[49]) or after('BASIS:',r[49]) or '').strip()
    low=full.lower()
    hits=sorted((low.find(kw),kw) for kw in ENERGY_KW if kw in low)
    found=[]                     # keep first-occurring; drop hits inside an already-kept span
    spans=[]                     # ('wind' inside 'offshore wind', etc.)
    for i,kw in hits:
        if any(s<=i<e for s,e in spans): continue
        found.append(kw); spans.append((i,i+len(kw)))
    found=found[:2]
    summary=' · '.join(kw.capitalize() for kw in found) if found else \
            (full.split(';')[0][:52].rstrip()+'…' if len(full.split(';')[0])>52 else full.split(';')[0])
    return rating, summary, full

CITY_RE = re.compile(r'([^(;]+?)\s*\(([\d.]+)\s*M\s*[—-]\s*([^)]+)\)')
def cities(r):
    out=[]
    for m in CITY_RE.finditer(str(r[46] or '')):
        name,pop,status=m.group(1).strip(),m.group(2),m.group(3).strip().lower()
        if 'defended' in status: lab,css,arr='Defended','st-defended','&#9660;'
        elif 'excluded' in status or 'contract' in status: lab,css,arr='Contracting','st-excluded','&#9660;'
        elif 'stress' in status: lab,css,arr='Stressed','st-stressed','&#9660;'
        elif 'viable' in status: lab,css,arr='Viable','st-viable',''
        else: lab,css,arr=status.title(),'st-viable',''
        out.append((name,f'{float(pop):g}',lab,css,arr))
    return out

def seasons_html(r, content):
    labels=content.get('season_labels')
    mo={'Winter':'Jul','Spring':'Oct','Summer':'Jan','Autumn':'Apr'} if str(r[65] or 'N').strip().upper()=='S' \
        else {'Winter':'Jan','Spring':'Apr','Summer':'Jul','Autumn':'Oct'}
    cells=[('Winter',16,18),('Spring',20,22),('Summer',24,26),('Autumn',28,30)]
    tds=''
    for i,(name,hiF,loF) in enumerate(cells):
        disp=labels[i] if labels else f'{mo[name]} · {name[:3]}'
        tds+=(f'<div class="seas"><div class="sn">{disp}</div>'
              f'<div class="hi">{r[hiF]}</div><div class="cc">{r[hiF+1]}</div><div class="hl">avg high</div>'
              f'<div class="lo">{r[loF]}</div><div class="cc">{r[loF+1]}</div><div class="hl">avg low</div></div>')
    return tds, (r[79] or '—'), (r[80] or '—')

def leadin(p, i):
    """Style the paragraph's opening as the drop-style lead-in. The copywriter marks
    it by writing the opening words in ALL CAPS; wrap that whole caps run. Fall back to
    the first 2-3 words for plain content-file prose."""
    if i != 0 or 'leadin' in p:
        return p
    words = p.split(' ')
    n = 0
    for w in words:
        letters = re.sub(r'[^A-Za-z]', '', w)
        if letters and letters == letters.upper() and letters != letters.lower():
            n += 1
        else:
            break
    if n >= 2:                                   # leading ALL-CAPS run (copywriter convention)
        run = ' '.join(words[:n])
        return f'<span class="leadin">{run}</span>' + p[len(run):]
    m = re.match(r'^([A-Z][\w’]*(?:\s+[\w’]+){0,2})\b', p)   # fallback: first 2-3 words
    return (f'<span class="leadin">{m.group(1)}</span>' + p[len(m.group(1)):]) if m else p

def render(r, c):
    P=cw_prose(r)
    pick=lambda k,d='': (P.get(k) or c.get(k) or d)
    peak=num(r[3]); now=num(r[4])
    peak_disp=f'{peak:g} M' if peak is not None else '—'
    now_disp=f'{now:g} M' if now is not None else '—'
    trend=str(r[5] or '').strip(); TU=trend.upper()
    if 'GROWTH' in TU:   arrow,adir,frame='&#9650;','up',   f'from {peak_disp} before the warming began'
    elif 'CONTRACT' in TU: arrow,adir,frame='&#9660;','down', f'from a {peak_disp} pre-Contraction peak'
    else:               arrow,adir,frame='&#9644;','flat', f'stable near its {peak_disp} level'
    delta_line=f'{frame} · <em>{trend}</em>'
    pop_row=(f'<tr><td class="k">Population 2060</td><td class="v"><strong>{now_disp}</strong> · '
             f'{frame} · trend <strong>{trend}</strong></td></tr>')
    wrate,wrisk=water(r); erate,edetail,efull=energy(r)
    efull_html=f' — {efull}' if efull else ''
    method=str(r[6] or '—').split('(')[0].strip()
    crops=str(r[42] or '').strip()
    production=method+(f' — {crops}' if crops else '')
    transport=after('TRANSPORT:',r[52]).rstrip(' .')
    ports=(after('PORTS 2060:',r[94]) or str(r[94] or '').strip()).rstrip(' .')
    iso=after('ISOLATION RISK:',r[52])
    cparts=[]
    if transport: cparts.append(transport)
    if ports: cparts.append(f'<strong>Ports 2060:</strong> {ports}')
    if iso: cparts.append(f'isolation risk {iso}')
    connectivity=' · '.join(cparts)
    seas_tds,exH,exL=seasons_html(r,c)
    hazards=str(r[34] or '').strip()
    rainfall=str(r[36] or '—').strip(); elev=str(r[37] or '—').strip()
    cityless = str(r[46] or '').strip().upper().startswith('CITYLESS')
    city_rows=''.join(f'<tr><td class="c">{n}</td><td class="p">{p} M</td>'
                      f'<td class="s {css}"><span class="dot"></span>{lab} {arr}</td></tr>'
                      for n,p,lab,css,arr in cities(r)[:6])
    cities_block = '' if (cityless or not city_rows) else CITIES_BLOCK.format(city_rows=city_rows)
    inflow=str(r[61] or '').strip()
    if inflow in ('','—','-','–','—','None','none','N/A','n/a','n.a.','0'): inflow=''
    inflow_html=(f'<tr><td class="k">Displaced inflow</td><td class="v">{inflow} — the nearest survivable '
                 f'direction, not a guaranteed seat; not added to the settled figure</td></tr>') if inflow else ''
    cdraw_col=str(r[77] or '').strip(); cdraw_html=''
    if cdraw_col:
        body=c.get('cultural_draw', re.sub(r'\s*TARGET:.*$','',cdraw_col,flags=re.S).strip())
        cdraw_html=f'<div class="callout draw"><div class="co-lab">Who this calls</div><div class="co-body">{body}</div></div>'
    legacy_col=str(r[78] or '').strip(); legacy_html=''
    if legacy_col and c.get('show_legacy',True):
        legacy_html=f'<div class="callout legacy"><div class="co-lab">Legacy &amp; knowledge</div><div class="co-body">{c.get("legacy_note",legacy_col)}</div></div>'
    lede=P['lede'] if P.get('lede') else c['lede']
    living=P['living'] if P.get('living') else c['living']
    free=P['free_hours'] if P.get('free_hours') else c['free_hours']
    lede_html=''.join(f'<p>{leadin(p,i)}</p>' for i,p in enumerate(lede))
    living_li=''.join(f'<li>{x}</li>' for x in living)
    free_li=''.join(f'<li class="alt">{x}</li>' for x in free)
    cw_caps=P.get('captions',[]); photos_html=''
    for i,(src,cap) in enumerate(c['photos']):
        caption=cw_caps[i] if i<len(cw_caps) and cw_caps[i] else cap
        photos_html+=f'<figure class="ph"><img src="{src}"><figcaption>{caption}</figcaption></figure>'
    hero = c.get('hero_svg') or hero_fpo(c.get('hero_note',''))
    cap0,cap1=c['hero_caption']; mx,my,mcoord=c['marker']
    mxf,myf=float(mx),float(my)
    marker_svg=(f'<circle cx="{mxf}" cy="{myf}" r="15" fill="none" stroke="#A6402B" stroke-width="1.3"/>'
                f'<circle cx="{mxf}" cy="{myf}" r="2.2" fill="#A6402B"/>'
                f'<line x1="{mxf-25}" y1="{myf}" x2="{mxf-7}" y2="{myf}" stroke="#A6402B" stroke-width="1"/>'
                f'<line x1="{mxf+7}" y1="{myf}" x2="{mxf+25}" y2="{myf}" stroke="#A6402B" stroke-width="1"/>'
                f'<line x1="{mxf}" y1="{myf-25}" x2="{mxf}" y2="{myf-7}" stroke="#A6402B" stroke-width="1"/>'
                f'<line x1="{mxf}" y1="{myf+7}" x2="{mxf}" y2="{myf+25}" stroke="#A6402B" stroke-width="1"/>')
    return TEMPLATE.format(
        rid=r[56], title=c['title'], eyebrow=c['eyebrow'], subhead=pick('subhead'),
        hero=hero, cap0=cap0, cap1=cap1,
        now=now_disp, arrow=arrow, adir=adir, delta_line=delta_line, pop_row=pop_row,
        wrate=wrate, wrisk=wrisk, erate=erate, edetail=edetail, efull=efull_html, exH=exH, exL=exL,
        production=production, connectivity=connectivity,
        wsrc=after('PRIMARY SOURCE:',r[9]) or rainfall, rainfall=rainfall, elev=elev, hazards=hazards,
        seasonality=str(r[15] or '').split('—')[0].strip(),
        cdraw=cdraw_html, legacy=legacy_html, inflow=inflow_html,
        lede=lede_html, holding=pick('holding'), living=living_li,
        table_prose=pick('table_prose'), free=free_li, ledger=pick('ledger'),
        cities_block=cities_block, seas=seas_tds, seasnote=c.get('season_note',''),
        marker_svg=marker_svg, mcoord=mcoord, worldmap=WORLDMAP,
        photos=photos_html, cta_big=pick('cta_big'), cta_small=pick('cta_small'))

def hero_fpo(note):
    body = note or 'Hero illustration TK — depiction note to be supplied.'
    return ('<div class="hero-fpo"><div class="fpo-tag">Hero illustration · FPO</div>'
            f'<div class="fpo-note">{body}</div></div>')

_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE=open(os.path.join(_DIR,'template_flow.html')).read()
CITIES_BLOCK=open(os.path.join(_DIR,'cities_block.html')).read()
WORLDMAP=open(os.path.join(_DIR,'worldmap_inner.svg')).read()

if __name__=='__main__':
    spec=importlib.util.spec_from_file_location('content', sys.argv[1])
    mod=importlib.util.module_from_spec(spec); spec.loader.exec_module(mod)
    c=mod.CONTENT; r=load_region(c['region_id'])
    out=sys.argv[2] if len(sys.argv)>2 else '/home/claude/gen/out.html'
    open(out,'w').write(render(r,c))
    src='CW columns' if cw_prose(r) else 'content-file fallback'
    print(f'[atlas_gen {GENERATOR_VERSION}] wrote {out} · {c["region_id"]} {str(r[1])[:34]} · prose from {src} · workbook {WB.split("/")[-1]}')
